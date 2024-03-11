# .\vkr\Scripts\activate

import tkinter as tk
from tkinter import tix
from tkinter.font import Font
from tkinter import filedialog

import numpy as np
import pandas as pd
import pickle
from sklearn.preprocessing import StandardScaler

import os, shutil, pathlib
import openpyxl as xlsx
from datetime import datetime

# paths
DOWNLOAD_PATH = str(pathlib.Path.home() / "Downloads")
SERVER_PATH = os.path.join(os.getcwd(), 'data')
TEMPLATE_FILE_PATH = os.path.join(SERVER_PATH, 'template.xlsx')
EXAMPLE_FILE_PATH = os.path.join(SERVER_PATH, 'example.xlsx')

# excel
EXCEL_A1_W1 = [f"{chr(x)}1" for x in range(ord('A'), ord('W') + 1)]
EXCEL_A2_W2 = [f"{chr(x)}2" for x in range(ord('A'), ord('W') + 1)]
EXCEL_HEADER = [
    '% Iron Feed', '% Silica Feed', 'Starch Flow', 'Amina Flow',
    'Ore Pulp Flow', 'Ore Pulp pH', 'Ore Pulp Density',
    'Flotation Column 01 Air Flow', 'Flotation Column 02 Air Flow',
    'Flotation Column 03 Air Flow', 'Flotation Column 04 Air Flow',
    'Flotation Column 05 Air Flow', 'Flotation Column 06 Air Flow',
    'Flotation Column 07 Air Flow',
    'Flotation Column 01 Level', 'Flotation Column 02 Level',
    'Flotation Column 03 Level', 'Flotation Column 04 Level',
    'Flotation Column 05 Level', 'Flotation Column 06 Level',
    'Flotation Column 07 Level',
    '% Iron Concentrate', '% Silica Concentrate'
]
EXCEL_DATA = [
    '55.2', '16.98', '3019.53', '557.434', '395.713', '10.0664',
    '1.74', '249.214', '253.235', '1.74', '295.096', '306.4',
    '250.225', '250.884', '457.396', '432.962', '424.954',
    '443.558', '502.255', '446.37', '523.344', '63.942', '2.89'
]

# upload models
with open(os.path.join(os.getcwd(), 'model', 'model.pkl'), 'rb') as f:
    MODEL = pickle.load(f)


def update_status(message: str, is_err: bool = False):
    label_status['fg'] = 'red' if is_err else 'black'
    label_status['text'] = f"ERROR: {message}." if is_err else f"SUCCESS: {message}."


def get_values():
    values = list()
    for obj in entry_objects:
        values.append(obj.get())

    return values


def delete_values(need_update_status: bool = True):
    for obj in entry_objects:
        if obj['state'] == 'readonly':
            obj['state'] = 'normal'
            obj.delete(0, tk.END)
            obj['state'] = 'readonly'
        else:
            obj.delete(0, tk.END)

    if need_update_status:
        update_status(f"Data reset")


def format_values(need_update_status: bool = True):
    curr_values = get_values()[:-2]  # don't need to format the readonly fields
    formatted_values = list()  # will have float or None
    trigger = False
    for count, value in enumerate(curr_values):
        value = str(value).strip()
        if value == '' or 'n/f' in value:
            formatted_values.append(None)
            # for check the empty fields
            if not trigger and count < len(curr_values) - 1:
                trigger = True
            continue

        value = value.replace(',', '.')
        try:
            value = round(float(value), 3)
            formatted_values.append(value)
        except ValueError:
            value += ' (n/f)'
            formatted_values.append(None)

        # update data
        obj = entry_objects[count]
        obj.delete(0, tk.END)
        obj.insert(0, str(value))

    if need_update_status:
        update_status(f"Data formatted")

    return formatted_values, trigger


def import_data_from_excel():
    file_path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])
    if not file_path:
        update_status(f"File not selected or corrupted", True)
        return

    # read data from file and update fields
    delete_values(False)
    work_sheet = xlsx.load_workbook(file_path).active
    for count, idx in enumerate(EXCEL_A2_W2):
        val = work_sheet[idx].value
        val = '' if val is None else val
        obj = entry_objects[count]
        if obj['state'] == 'readonly':
            obj['state'] = 'normal'
            obj.insert(0, val)
            obj['state'] = 'readonly'
        else:
            obj.insert(0, val)

    update_status(f"Data filled from excel file")


def calculate():
    formatted_values, trigger = format_values(False)
    if trigger:
        update_status(f"Data not formatted or skipped", True)
        return

    if len(formatted_values) != len(EXCEL_HEADER[:-2]):
        update_status(f"Got incorrect data", True)
        return

    # data processes and model prediction
    x_data = pd.DataFrame(data=dict(zip(EXCEL_HEADER[:-2], formatted_values)), index=[0]).astype(np.float64)
    scale = StandardScaler()
    x_data = pd.DataFrame(scale.fit_transform(x_data))
    iron, silica = MODEL.predict(x_data)[0]

    if iron is None or silica is None:
        update_status(f"Got incorrect data", True)
        return

    # update iron
    iron = round(float(iron), 3)
    input_iron_concentrate['state'] = 'normal'
    input_iron_concentrate.delete(0, tk.END)
    input_iron_concentrate.insert(0, str(iron))
    input_iron_concentrate['state'] = 'readonly'

    # update silica
    silica = round(float(silica), 3)
    input_silica_concentrate['state'] = 'normal'
    input_silica_concentrate.delete(0, tk.END)
    input_silica_concentrate.insert(0, str(silica))
    input_silica_concentrate['state'] = 'readonly'

    update_status(f"Computation complete")


def download_data_to_excel():
    data = get_values()
    work_book = xlsx.Workbook()
    work_sheet = work_book.active
    work_sheet.title = "Result"
    # fill first row (headers)
    for count, idx in enumerate(EXCEL_A1_W1):
        work_sheet[idx] = EXCEL_HEADER[count]
    # fill second row (data)
    for count, idx in enumerate(EXCEL_A2_W2):
        work_sheet[idx] = '' if data[count] is None else data[count]

    # example file name: "result_20012024_134218.xlsx"
    file_name = f"result_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx"
    work_book.save(os.path.join(DOWNLOAD_PATH, file_name))
    update_status(f"File '{file_name}' downloaded")


def download_template_or_example_file(file_name: str = 'template'):
    file_path = EXAMPLE_FILE_PATH if file_name == 'example' else TEMPLATE_FILE_PATH
    # create file on server
    if not os.path.isfile(file_path):
        work_book = xlsx.Workbook()
        work_sheet = work_book.active
        work_sheet.title = file_name
        # fill first row (headers)
        for count, idx in enumerate(EXCEL_A1_W1):
            work_sheet[idx] = EXCEL_HEADER[count]
        if file_name == 'example':
            # fill second row (data)
            for count, idx in enumerate(EXCEL_A2_W2):
                if count > len(EXCEL_DATA):
                    break
                work_sheet[idx] = '' if EXCEL_DATA[count] is None else EXCEL_DATA[count]

        work_book.save(file_path)

    shutil.copy2(file_path, DOWNLOAD_PATH)
    update_status(f"File '{file_name}.xlsx' downloaded")


# program

root = tix.Tk()
root.title('Main window')
root.resizable(False, False)
root['bg'] = 'light blue'

# menu > file
menubar = tk.Menu(root)
file_menu = tk.Menu(menubar, tearoff=0)
file_menu.add_command(label="Import", command=import_data_from_excel)
file_menu.add_command(label="Format", command=format_values)
file_menu.add_command(label="Reset", command=delete_values)
file_menu.add_separator()
file_menu.add_command(label="Exit", command=root.quit)
menubar.add_cascade(label="File", menu=file_menu)
# menu > download
download_menu = tk.Menu(menubar, tearoff=0)
download_menu.add_command(label="Result", command=download_data_to_excel)
download_menu.add_separator()
download_menu.add_command(label="Template", command=lambda: download_template_or_example_file('template'))
download_menu.add_command(label="Example", command=lambda: download_template_or_example_file('example'))
menubar.add_cascade(label="Download", menu=download_menu)
# add menu to window
root['menu'] = menubar

COMMON_STYLE = {'master': root, 'font': Font(family='Arial', size=14)}
STYLE_LABEL = {**COMMON_STYLE, 'bg': 'light blue'}
STYLE_INPUT = {**COMMON_STYLE, 'bg': 'snow', 'width': 10, 'readonlybackground': 'gray85'}
STYLE_BUTTON = {**COMMON_STYLE, 'bg': 'SkyBlue2', 'width': 10}
PADDINGS_INPUT = {'padx': 5, 'pady': 5, 'ipadx': 2, 'ipady': 2}
PADDINGS_BUTTON = {'padx': 10, 'pady': 20, 'ipadx': 2, 'ipady': 2}

entry_objects = []

# column 1

label_iron_feed = tk.Label(**STYLE_LABEL, text='Iron Feed, %:')
input_iron_feed = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_iron_feed)
label_iron_feed.grid(**PADDINGS_INPUT, row=0, column=0)
input_iron_feed.grid(**PADDINGS_INPUT, row=0, column=1)

label_silica_feed = tk.Label(**STYLE_LABEL, text='Silica Feed, %:')
input_silica_feed = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_silica_feed)
label_silica_feed.grid(**PADDINGS_INPUT, row=1, column=0)
input_silica_feed.grid(**PADDINGS_INPUT, row=1, column=1)

label_starch_flow = tk.Label(**STYLE_LABEL, text='Starch Flow:')
input_starch_flow = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_starch_flow)
label_starch_flow.grid(**PADDINGS_INPUT, row=2, column=0)
input_starch_flow.grid(**PADDINGS_INPUT, row=2, column=1)

label_amina_flow = tk.Label(**STYLE_LABEL, text='Amina Flow:')
input_amina_flow = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_amina_flow)
label_amina_flow.grid(**PADDINGS_INPUT, row=3, column=0)
input_amina_flow.grid(**PADDINGS_INPUT, row=3, column=1)

label_ore_pulp_flow = tk.Label(**STYLE_LABEL, text='Ore Pulp Flow:')
input_ore_pulp_flow = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_ore_pulp_flow)
label_ore_pulp_flow.grid(**PADDINGS_INPUT, row=4, column=0)
input_ore_pulp_flow.grid(**PADDINGS_INPUT, row=4, column=1)

label_ore_pulp_ph = tk.Label(**STYLE_LABEL, text='Ore Pulp pH:')
input_ore_pulp_ph = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_ore_pulp_ph)
label_ore_pulp_ph.grid(**PADDINGS_INPUT, row=5, column=0)
input_ore_pulp_ph.grid(**PADDINGS_INPUT, row=5, column=1)

label_ore_pulp_density = tk.Label(**STYLE_LABEL, text='Ore Pulp Density:')
input_ore_pulp_density = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_ore_pulp_density)
label_ore_pulp_density.grid(**PADDINGS_INPUT, row=6, column=0)
input_ore_pulp_density.grid(**PADDINGS_INPUT, row=6, column=1)

# column 2

label_air_flow_1 = tk.Label(**STYLE_LABEL, text='Flotation Col 01\nAir Flow:')
input_air_flow_1 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_air_flow_1)
label_air_flow_1.grid(**PADDINGS_INPUT, row=0, column=2)
input_air_flow_1.grid(**PADDINGS_INPUT, row=0, column=3)

label_air_flow_2 = tk.Label(**STYLE_LABEL, text='Flotation Col 02\nAir Flow:')
input_air_flow_2 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_air_flow_2)
label_air_flow_2.grid(**PADDINGS_INPUT, row=1, column=2)
input_air_flow_2.grid(**PADDINGS_INPUT, row=1, column=3)

label_air_flow_3 = tk.Label(**STYLE_LABEL, text='Flotation Col 03\nAir Flow:')
input_air_flow_3 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_air_flow_3)
label_air_flow_3.grid(**PADDINGS_INPUT, row=2, column=2)
input_air_flow_3.grid(**PADDINGS_INPUT, row=2, column=3)

label_air_flow_4 = tk.Label(**STYLE_LABEL, text='Flotation Col 04\nAir Flow:')
input_air_flow_4 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_air_flow_4)
label_air_flow_4.grid(**PADDINGS_INPUT, row=3, column=2)
input_air_flow_4.grid(**PADDINGS_INPUT, row=3, column=3)

label_air_flow_5 = tk.Label(**STYLE_LABEL, text='Flotation Col 05\nAir Flow:')
input_air_flow_5 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_air_flow_5)
label_air_flow_5.grid(**PADDINGS_INPUT, row=4, column=2)
input_air_flow_5.grid(**PADDINGS_INPUT, row=4, column=3)

label_air_flow_6 = tk.Label(**STYLE_LABEL, text='Flotation Col 06\nAir Flow:')
input_air_flow_6 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_air_flow_6)
label_air_flow_6.grid(**PADDINGS_INPUT, row=5, column=2)
input_air_flow_6.grid(**PADDINGS_INPUT, row=5, column=3)

label_air_flow_7 = tk.Label(**STYLE_LABEL, text='Flotation Col 07\nAir Flow:')
input_air_flow_7 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_air_flow_7)
label_air_flow_7.grid(**PADDINGS_INPUT, row=6, column=2)
input_air_flow_7.grid(**PADDINGS_INPUT, row=6, column=3)

# column 3

label_level_1 = tk.Label(**STYLE_LABEL, text='Flotation Col 01\nLevel:')
input_level_1 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_level_1)
label_level_1.grid(**PADDINGS_INPUT, row=0, column=4)
input_level_1.grid(**PADDINGS_INPUT, row=0, column=5)

label_level_2 = tk.Label(**STYLE_LABEL, text='Flotation Col 02\nLevel:')
input_level_2 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_level_2)
label_level_2.grid(**PADDINGS_INPUT, row=1, column=4)
input_level_2.grid(**PADDINGS_INPUT, row=1, column=5)

label_level_3 = tk.Label(**STYLE_LABEL, text='Flotation Col 03\nLevel:')
input_level_3 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_level_3)
label_level_3.grid(**PADDINGS_INPUT, row=2, column=4)
input_level_3.grid(**PADDINGS_INPUT, row=2, column=5)

label_level_4 = tk.Label(**STYLE_LABEL, text='Flotation Col 04\nLevel:')
input_level_4 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_level_4)
label_level_4.grid(**PADDINGS_INPUT, row=3, column=4)
input_level_4.grid(**PADDINGS_INPUT, row=3, column=5)

label_level_5 = tk.Label(**STYLE_LABEL, text='Flotation Col 05\nLevel:')
input_level_5 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_level_5)
label_level_5.grid(**PADDINGS_INPUT, row=4, column=4)
input_level_5.grid(**PADDINGS_INPUT, row=4, column=5)

label_level_6 = tk.Label(**STYLE_LABEL, text='Flotation Col 06\nLevel:')
input_level_6 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_level_6)
label_level_6.grid(**PADDINGS_INPUT, row=5, column=4)
input_level_6.grid(**PADDINGS_INPUT, row=5, column=5)

label_level_7 = tk.Label(**STYLE_LABEL, text='Flotation Col 07\nLevel:')
input_level_7 = tk.Entry(**STYLE_INPUT)
entry_objects.append(input_level_7)
label_level_7.grid(**PADDINGS_INPUT, row=6, column=4)
input_level_7.grid(**PADDINGS_INPUT, row=6, column=5)

# last row (readonly fields)

hint_auto_calc = tix.Balloon(root)

label_iron_concentrate = tk.Label(**STYLE_LABEL, text='Iron Concentrate, %:')
input_iron_concentrate = tk.Entry(**STYLE_INPUT, state='readonly')
entry_objects.append(input_iron_concentrate)
label_iron_concentrate.grid(**PADDINGS_INPUT, row=7, column=0)
input_iron_concentrate.grid(**PADDINGS_INPUT, row=7, column=1)
hint_auto_calc.bind_widget(input_iron_concentrate, balloonmsg='Automatic calculation')

label_silica_concentrate = tk.Label(**STYLE_LABEL, text='Silica Concentrate, %:')
input_silica_concentrate = tk.Entry(**STYLE_INPUT, state='readonly')
entry_objects.append(input_silica_concentrate)
label_silica_concentrate.grid(**PADDINGS_INPUT, row=7, column=2)
input_silica_concentrate.grid(**PADDINGS_INPUT, row=7, column=3)
hint_auto_calc.bind_widget(input_silica_concentrate, balloonmsg='Automatic calculation')

# calculate button
button = tk.Button(**STYLE_BUTTON, text='Calculate', command=calculate)
button.grid(**PADDINGS_BUTTON, row=7, column=4, columnspan=2)

# status label
label_status = tk.Label(**STYLE_LABEL, text='')
label_status.grid(**PADDINGS_INPUT, row=9, column=0, columnspan=5, sticky='W')

root.mainloop()
